#!/usr/bin/env python3
"""
Export ordinance/amendment questionnaire answers to an Excel workbook.

Inputs per run:
  - request dir: mapping_shard*.jsonl (from scripts/export_ordinance_questionnaire_batch_requests.py)
  - results dir: normalized.jsonl + invalid_rows.jsonl (from scripts/normalize_ordinance_questionnaire_openai_batch_results.py)

This produces a derived artifact (XLSX) intended for human inspection and ad-hoc analysis.
The normalized.jsonl remains the source of truth for downstream compute.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from os.path import expanduser
from pathlib import Path
from typing import Any


def _eprint(msg: str) -> None:
    print(msg, flush=True)


def _json_dumps_compact(v: Any) -> str:
    return json.dumps(v, ensure_ascii=False, separators=(",", ":"))


def _page_id_to_slug(page_id: Any) -> str | None:
    """Best-effort parse of a newspaper slug from a page_id.

    Expected format: <slug>-<mon>-<dd>-<yyyy>-p-<page>
    We treat the slug as everything before the last 5 dash-separated parts.
    """

    if not isinstance(page_id, str):
        return None
    pid = page_id.strip()
    if not pid:
        return None
    parts = pid.split("-")
    if len(parts) < 6:
        return None
    slug = "-".join(parts[:-5]).strip()
    return slug or None


@dataclass(frozen=True)
class Question:
    id: str
    question_type: str
    short_question: str | None
    full_question: str | None
    possible_answers: list[str] | None


@dataclass
class _QuestionStats:
    total: int = 0
    answered: int = 0
    conf_sum_answered: float = 0.0
    conf_min_answered: float | None = None
    conf_max_answered: float | None = None
    evidence_present_answered: int = 0
    unit_present_answered: int = 0
    true_count: int = 0
    false_count: int = 0
    categorical_counts: Counter[str] | None = None
    numeric_values: list[float] | None = None
    unit_counts: Counter[str] | None = None


def _load_questions_xlsx(*, xlsx_path: Path, processed_sheet: str, input_info_sheet: str) -> list[Question]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise SystemExit("Missing dependency: openpyxl (pip install openpyxl)") from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if processed_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {processed_sheet!r} (have {wb.sheetnames})")
    if input_info_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {input_info_sheet!r} (have {wb.sheetnames})")

    ws_proc = wb[processed_sheet]
    ws_in = wb[input_info_sheet]

    def _hdr_map(ws) -> dict[str, int]:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        out: dict[str, int] = {}
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            key = str(h).strip()
            if key:
                out[key] = i
        return out

    proc_idx = _hdr_map(ws_proc)
    in_idx = _hdr_map(ws_in)

    required_proc = {"ID", "Include", "Question Type"}
    missing_proc = sorted(required_proc - set(proc_idx))
    if missing_proc:
        raise SystemExit(f"Processed sheet missing headers: {missing_proc}")

    possible_by_id: dict[str, list[str]] = {}
    if "ID" in in_idx and "Possible Answers" in in_idx:
        for r in range(2, ws_in.max_row + 1):
            qid_raw = ws_in.cell(r, in_idx["ID"]).value
            if qid_raw is None:
                continue
            qid = str(qid_raw).strip()
            if not qid:
                continue
            poss_raw = ws_in.cell(r, in_idx["Possible Answers"]).value
            if poss_raw is None:
                continue
            poss_s = str(poss_raw).strip()
            if not poss_s:
                continue
            opts = [p.strip() for p in poss_s.split(";") if p.strip()]
            if opts:
                possible_by_id[qid] = opts

    questions: list[Question] = []
    for r in range(2, ws_proc.max_row + 1):
        include_raw = ws_proc.cell(r, proc_idx["Include"]).value
        if str(include_raw).strip().lower() != "yes":
            continue

        qid_raw = ws_proc.cell(r, proc_idx["ID"]).value
        if qid_raw is None:
            continue
        qid = str(qid_raw).strip()
        if not qid:
            continue

        qtype_raw = ws_proc.cell(r, proc_idx["Question Type"]).value
        qtype = str(qtype_raw).strip() if qtype_raw is not None else ""
        if qtype not in {"Binary", "Categorical", "Numerical", "Continuous"}:
            raise SystemExit(f"Unsupported Question Type {qtype!r} for ID={qid} (row {r})")

        short_q = None
        if "Short Question" in proc_idx:
            v = ws_proc.cell(r, proc_idx["Short Question"]).value
            short_q = str(v).strip() if v is not None and str(v).strip() else None
        full_q = None
        if "Full Question" in proc_idx:
            v = ws_proc.cell(r, proc_idx["Full Question"]).value
            full_q = str(v).strip() if v is not None and str(v).strip() else None

        questions.append(
            Question(
                id=qid,
                question_type=qtype,
                short_question=short_q,
                full_question=full_q,
                possible_answers=possible_by_id.get(qid),
            )
        )

    if not questions:
        raise SystemExit("No included questions found (Include == Yes)")
    return questions


def _read_mapping_by_id(request_dir: Path) -> dict[str, dict[str, Any]]:
    mapping_by_id: dict[str, dict[str, Any]] = {}
    paths = sorted(request_dir.glob("mapping_shard*.jsonl"))
    if not paths:
        raise SystemExit(f"No mapping_shard*.jsonl found in {request_dir}")
    for mp in paths:
        for raw in mp.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            cid = obj.get("id") or obj.get("custom_id")
            if isinstance(cid, str) and cid:
                mapping_by_id[cid] = obj
    return mapping_by_id


def _iter_jsonl(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            yield json.loads(line)


def _flatten_bbox(bbox: Any) -> tuple[int | None, int | None, int | None, int | None]:
    if not isinstance(bbox, dict):
        return None, None, None, None
    x0 = bbox.get("x0")
    y0 = bbox.get("y0")
    x1 = bbox.get("x1")
    y1 = bbox.get("y1")
    return (
        int(x0) if isinstance(x0, int) else None,
        int(y0) if isinstance(y0, int) else None,
        int(x1) if isinstance(x1, int) else None,
        int(y1) if isinstance(y1, int) else None,
    )


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Export ordinance questionnaire answers to a single XLSX workbook.")
    ap.add_argument(
        "--run",
        action="append",
        nargs=2,
        metavar=("REQUEST_DIR", "RESULTS_DIR"),
        help=(
            "A (request_dir, results_dir) pair. Repeatable. "
            "request_dir must contain mapping_shard*.jsonl; results_dir must contain normalized.jsonl + invalid_rows.jsonl."
        ),
        required=True,
    )
    ap.add_argument("--questions-xlsx", required=True, help="Path to Questions.xlsx")
    ap.add_argument("--questions-processed-sheet", default="Processed Info", help="Processed sheet name")
    ap.add_argument("--questions-input-info-sheet", default="Input Info", help="Input Info sheet name")
    ap.add_argument(
        "--out-xlsx",
        required=True,
        help="Path to write the output workbook to",
    )
    ap.add_argument(
        "--include-null-answers",
        action="store_true",
        help="Include rows/columns for questions where answer is null (default: include; this flag is kept for clarity).",
    )
    ap.add_argument(
        "--omit-evidence",
        action="store_true",
        help="Do not include evidence columns (keeps the workbook smaller).",
    )
    ap.add_argument(
        "--evidence-max-chars",
        type=int,
        default=0,
        help="If >0, truncate evidence strings to this many characters.",
    )
    ap.add_argument(
        "--on-duplicate-custom-id",
        choices=["error", "keep_first", "keep_last"],
        default="error",
        help="What to do if the same custom_id appears in multiple runs.",
    )
    ap.add_argument(
        "--max-boxes",
        type=int,
        default=0,
        help="Debug: only export the first N box rows (0 = all).",
    )
    return ap.parse_args()


def main() -> None:
    args = _parse_args()

    runs: list[tuple[Path, Path]] = []
    for req_s, res_s in args.run:
        req = Path(expanduser(req_s)).resolve()
        res = Path(expanduser(res_s)).resolve()
        if not req.is_dir():
            raise SystemExit(f"request_dir not found: {req}")
        if not res.is_dir():
            raise SystemExit(f"results_dir not found: {res}")
        runs.append((req, res))

    questions_xlsx = Path(expanduser(args.questions_xlsx)).resolve()
    if not questions_xlsx.is_file():
        raise SystemExit(f"Questions.xlsx not found: {questions_xlsx}")

    questions = _load_questions_xlsx(
        xlsx_path=questions_xlsx,
        processed_sheet=str(args.questions_processed_sheet),
        input_info_sheet=str(args.questions_input_info_sheet),
    )
    qids = [q.id for q in questions]
    q_by_id = {q.id: q for q in questions}

    out_xlsx = Path(expanduser(args.out_xlsx)).resolve()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise SystemExit("Missing dependency: openpyxl (pip install openpyxl)") from exc

    wb = openpyxl.Workbook(write_only=True)

    ws_runs = wb.create_sheet("runs")
    ws_questions = wb.create_sheet("questions")
    ws_invalid = wb.create_sheet("invalid_rows")
    ws_answers = wb.create_sheet("answers_wide")
    ws_overall = wb.create_sheet("overall_stats")
    ws_q_summary = wb.create_sheet("question_summary")
    ws_value_counts = wb.create_sheet("question_value_counts")

    ws_runs.append(
        [
            "run_name",
            "request_dir",
            "results_dir",
            "normalized_path",
            "invalid_path",
            "normalized_rows",
            "invalid_rows",
        ]
    )
    ws_questions.append(["question_id", "question_type", "short_question", "full_question", "possible_answers"])
    for qid in qids:
        q = q_by_id[qid]
        ws_questions.append(
            [
                q.id,
                q.question_type,
                q.short_question,
                q.full_question,
                ";".join(q.possible_answers) if q.possible_answers else None,
            ]
        )

    ws_invalid.append(["run_name", "custom_id", "page_id", "slug", "reason", "error"])

    # Build the answers_wide header.
    header = [
        "run_name",
        "custom_id",
        "slug",
        "page_id",
        "box_id",
        "classifier_label",
        "classifier_confidence",
        "cls",
        "bbox_x0",
        "bbox_y0",
        "bbox_x1",
        "bbox_y1",
        "source_page_path",
        "issues",
        "notes",
    ]
    for qid in qids:
        header.append(f"{qid}__answer")
        header.append(f"{qid}__unit")
        header.append(f"{qid}__confidence")
        if not bool(args.omit_evidence):
            header.append(f"{qid}__evidence")
    ws_answers.append(header)

    # Stats (across all included runs after de-dupe).
    boxes_with_any_non_null_answer = 0
    q_stats: dict[str, _QuestionStats] = {}
    for qid in qids:
        q = q_by_id[qid]
        st = _QuestionStats()
        if q.question_type in {"Categorical", "Binary"}:
            st.categorical_counts = Counter()
        if q.question_type in {"Numerical", "Continuous"}:
            st.numeric_values = []
        st.unit_counts = Counter()
        q_stats[qid] = st

    seen_custom_ids: dict[str, str] = {}  # custom_id -> run_name

    def handle_duplicate(custom_id: str, *, run_name: str) -> bool:
        """Return True if we should skip this record."""
        if custom_id not in seen_custom_ids:
            seen_custom_ids[custom_id] = run_name
            return False
        prev = seen_custom_ids[custom_id]
        mode = str(args.on_duplicate_custom_id)
        if mode == "error":
            raise SystemExit(f"Duplicate custom_id across runs: {custom_id} (runs: {prev}, {run_name})")
        if mode == "keep_first":
            return True
        if mode == "keep_last":
            # Keep last means: allow overwrite. For XLSX streaming, we can't delete prior rows.
            # So we fail loudly instead of silently producing ambiguous output.
            raise SystemExit(
                "on_duplicate_custom_id=keep_last is not supported for XLSX streaming; "
                f"duplicate custom_id: {custom_id} (runs: {prev}, {run_name})"
            )
        raise SystemExit(f"Unknown on_duplicate_custom_id mode: {mode}")

    max_boxes = int(args.max_boxes or 0)
    boxes_written = 0

    for req_dir, res_dir in runs:
        run_name = res_dir.name
        mapping_by_id = _read_mapping_by_id(req_dir)

        norm_path = res_dir / "normalized.jsonl"
        bad_path = res_dir / "invalid_rows.jsonl"
        if not norm_path.is_file():
            raise SystemExit(f"normalized.jsonl not found: {norm_path}")
        if not bad_path.is_file():
            raise SystemExit(f"invalid_rows.jsonl not found: {bad_path}")

        norm_rows = 0
        bad_rows = 0
        for bad in _iter_jsonl(bad_path):
            bad_rows += 1
            cid = bad.get("custom_id")
            mapping = mapping_by_id.get(cid) if isinstance(cid, str) else None
            page_id = mapping.get("page_id") if isinstance(mapping, dict) else None
            slug = _page_id_to_slug(page_id)
            ws_invalid.append([run_name, cid, page_id, slug, bad.get("reason"), bad.get("error")])

        for row in _iter_jsonl(norm_path):
            if max_boxes and boxes_written >= max_boxes:
                break
            norm_rows += 1
            custom_id = row.get("custom_id")
            if not isinstance(custom_id, str) or not custom_id:
                continue
            if handle_duplicate(custom_id, run_name=run_name):
                continue

            mapping = mapping_by_id.get(custom_id) or {}
            page_id = mapping.get("page_id") or row.get("page_id")
            slug = _page_id_to_slug(page_id)
            box_id = mapping.get("box_id") or row.get("box_id")
            classifier_label = mapping.get("classifier_label")
            classifier_confidence = mapping.get("classifier_confidence")

            bbox = mapping.get("bbox") if mapping.get("bbox") is not None else row.get("bbox")
            x0, y0, x1, y1 = _flatten_bbox(bbox)

            cls = mapping.get("cls") or row.get("cls")
            source_page_path = mapping.get("page_path") or row.get("source_page_path")
            issues = row.get("issues")
            issues_s = ";".join(issues) if isinstance(issues, list) else None
            normalized = row.get("normalized")
            if not isinstance(normalized, dict):
                raise SystemExit(f"Missing normalized object for custom_id={custom_id} in {norm_path}")
            notes = normalized.get("notes")
            notes_s = str(notes).strip() if isinstance(notes, str) and notes.strip() else None

            answers_by_id = normalized.get("answers_by_id")
            if not isinstance(answers_by_id, dict):
                raise SystemExit(f"Missing answers_by_id for custom_id={custom_id} in {norm_path}")

            any_non_null = False

            out_row: list[Any] = [
                run_name,
                custom_id,
                slug,
                page_id,
                box_id,
                classifier_label,
                classifier_confidence,
                cls,
                x0,
                y0,
                x1,
                y1,
                source_page_path,
                issues_s,
                notes_s,
            ]

            for qid in qids:
                ans = answers_by_id.get(qid)
                if not isinstance(ans, dict):
                    raise SystemExit(f"answers_by_id missing dict for qid={qid} custom_id={custom_id}")
                answer_val_raw = ans.get("answer")
                unit_val_raw = ans.get("unit")
                conf_val_raw = ans.get("confidence")
                ev_val_raw = ans.get("evidence")

                st = q_stats[qid]
                st.total += 1
                if answer_val_raw is not None:
                    any_non_null = True
                    st.answered += 1

                    if isinstance(conf_val_raw, (int, float)):
                        c = float(conf_val_raw)
                        st.conf_sum_answered += c
                        if st.conf_min_answered is None or c < st.conf_min_answered:
                            st.conf_min_answered = c
                        if st.conf_max_answered is None or c > st.conf_max_answered:
                            st.conf_max_answered = c

                    if isinstance(ev_val_raw, str) and ev_val_raw.strip():
                        st.evidence_present_answered += 1

                    if isinstance(unit_val_raw, str) and unit_val_raw.strip():
                        st.unit_present_answered += 1
                        if st.unit_counts is not None:
                            st.unit_counts[unit_val_raw.strip()] += 1

                    q = q_by_id[qid]
                    if q.question_type == "Binary":
                        if answer_val_raw is True:
                            st.true_count += 1
                        elif answer_val_raw is False:
                            st.false_count += 1
                    elif q.question_type == "Categorical":
                        if isinstance(answer_val_raw, str) and st.categorical_counts is not None:
                            st.categorical_counts[answer_val_raw] += 1
                    elif q.question_type in {"Numerical", "Continuous"}:
                        if isinstance(answer_val_raw, (int, float)) and st.numeric_values is not None:
                            st.numeric_values.append(float(answer_val_raw))

                # Values for Excel (may need JSON serialization)
                # Presentation rule: if the answer is null, leave ALL per-question cells blank in Excel.
                # The normalized JSON keeps confidence=0.0 for null answers (schema invariant),
                # but Excel is a human-facing artifact and "0.0" is misleading noise here.
                if answer_val_raw is None:
                    answer_val = None
                    unit_val = None
                    conf_val = None
                    ev_val = None
                else:
                    answer_val = answer_val_raw
                    unit_val = unit_val_raw
                    conf_val = conf_val_raw
                    ev_val = ev_val_raw

                # openpyxl can handle bool/int/float/str/None; serialize other JSON types.
                if isinstance(answer_val, (dict, list)):
                    answer_val = _json_dumps_compact(answer_val)
                if isinstance(unit_val, (dict, list)):
                    unit_val = _json_dumps_compact(unit_val)
                if isinstance(ev_val, (dict, list)):
                    ev_val = _json_dumps_compact(ev_val)

                if isinstance(ev_val, str) and args.evidence_max_chars and int(args.evidence_max_chars) > 0:
                    ev_val = ev_val[: int(args.evidence_max_chars)]

                out_row.append(answer_val)
                out_row.append(unit_val)
                out_row.append(conf_val)
                if not bool(args.omit_evidence):
                    out_row.append(ev_val)

            ws_answers.append(out_row)
            boxes_written += 1
            if any_non_null:
                boxes_with_any_non_null_answer += 1

        ws_runs.append(
            [
                run_name,
                str(req_dir),
                str(res_dir),
                str(norm_path),
                str(bad_path),
                norm_rows,
                bad_rows,
            ]
        )

    # Overall stats
    ws_overall.append(["metric", "value"])
    ws_overall.append(["boxes_exported", boxes_written])
    ws_overall.append(["boxes_with_any_non_null_answer", boxes_with_any_non_null_answer])
    ws_overall.append(
        [
            "boxes_with_any_non_null_answer_pct",
            (boxes_with_any_non_null_answer / boxes_written) if boxes_written else 0.0,
        ]
    )

    # Per-question summary
    ws_q_summary.append(
        [
            "question_id",
            "question_type",
            "short_question",
            "full_question",
            "possible_answers",
            "total_boxes",
            "answered_boxes",
            "answered_pct",
            "avg_conf_answered",
            "min_conf_answered",
            "max_conf_answered",
            "evidence_present_answered",
            "evidence_present_answered_pct",
            "unit_present_answered",
            "unit_present_answered_pct",
            "binary_true_count",
            "binary_false_count",
            "categorical_distinct_answers",
            "categorical_top_answer",
            "categorical_top_answer_count",
            "categorical_top_answer_pct_answered",
            "numeric_n",
            "numeric_min",
            "numeric_p10",
            "numeric_p50",
            "numeric_mean",
            "numeric_p90",
            "numeric_max",
            "units_top",
        ]
    )

    def _quantile(vals_sorted: list[float], q: float) -> float | None:
        if not vals_sorted:
            return None
        if q <= 0:
            return vals_sorted[0]
        if q >= 1:
            return vals_sorted[-1]
        idx = int(round((len(vals_sorted) - 1) * q))
        return vals_sorted[idx]

    for qid in qids:
        q = q_by_id[qid]
        st = q_stats[qid]
        total = st.total
        answered = st.answered
        answered_pct = (answered / total) if total else 0.0
        avg_conf = (st.conf_sum_answered / answered) if answered else None
        evidence_pct = (st.evidence_present_answered / answered) if answered else None
        unit_pct = (st.unit_present_answered / answered) if answered else None

        top_ans = None
        top_count = None
        top_pct = None
        distinct = None
        if q.question_type == "Categorical" and st.categorical_counts is not None:
            distinct = len(st.categorical_counts)
            if st.categorical_counts:
                top_ans, top_count = st.categorical_counts.most_common(1)[0]
                top_pct = (top_count / answered) if answered else None

        numeric_vals = st.numeric_values or []
        numeric_sorted = sorted(numeric_vals)
        numeric_n = len(numeric_vals)
        numeric_min = numeric_sorted[0] if numeric_sorted else None
        numeric_max = numeric_sorted[-1] if numeric_sorted else None
        numeric_mean = (sum(numeric_vals) / numeric_n) if numeric_n else None
        numeric_p10 = _quantile(numeric_sorted, 0.10)
        numeric_p50 = _quantile(numeric_sorted, 0.50)
        numeric_p90 = _quantile(numeric_sorted, 0.90)

        units_top = None
        if st.unit_counts is not None and st.unit_counts:
            units_top = _json_dumps_compact(dict(st.unit_counts.most_common(10)))

        ws_q_summary.append(
            [
                qid,
                q.question_type,
                q.short_question,
                q.full_question,
                ";".join(q.possible_answers) if q.possible_answers else None,
                total,
                answered,
                answered_pct,
                avg_conf,
                st.conf_min_answered,
                st.conf_max_answered,
                st.evidence_present_answered,
                evidence_pct,
                st.unit_present_answered,
                unit_pct,
                st.true_count if q.question_type == "Binary" else None,
                st.false_count if q.question_type == "Binary" else None,
                distinct,
                top_ans,
                top_count,
                top_pct,
                numeric_n if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_min if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_p10 if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_p50 if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_mean if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_p90 if q.question_type in {"Numerical", "Continuous"} else None,
                numeric_max if q.question_type in {"Numerical", "Continuous"} else None,
                units_top,
            ]
        )

    # Value counts (Binary + Categorical)
    ws_value_counts.append(["question_id", "question_type", "answer_value", "count", "pct_of_total", "pct_of_answered"])
    for qid in qids:
        q = q_by_id[qid]
        st = q_stats[qid]
        total = st.total
        answered = st.answered
        if q.question_type == "Binary":
            for val, cnt in [("true", st.true_count), ("false", st.false_count)]:
                if cnt <= 0:
                    continue
                ws_value_counts.append(
                    [qid, q.question_type, val, cnt, (cnt / total) if total else 0.0, (cnt / answered) if answered else 0.0]
                )
        elif q.question_type == "Categorical" and st.categorical_counts is not None:
            for val, cnt in st.categorical_counts.most_common():
                ws_value_counts.append(
                    [qid, q.question_type, val, cnt, (cnt / total) if total else 0.0, (cnt / answered) if answered else 0.0]
                )

    wb.save(out_xlsx)
    _eprint(f"wrote_xlsx\t{out_xlsx}")
    _eprint(f"runs\t{len(runs)}")
    _eprint(f"boxes_written\t{boxes_written}")
    _eprint(f"boxes_with_any_non_null_answer\t{boxes_with_any_non_null_answer}")


if __name__ == "__main__":
    main()
